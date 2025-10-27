# ValentinaPorSiempre.py
import streamlit as st
import pandas as pd
from datetime import date, datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path
import base64
from supabase import create_client
from dotenv import load_dotenv
import os

# ==========================================================
#                 LOAD ENVIRONMENT VARIABLES
# ==========================================================
load_dotenv()
SUPABASE_URL = os.getenv("SUPABASE_URL", "https://uumezwowrtumbonsotyc.supabase.co")
SUPABASE_KEY = os.getenv(
    "SUPABASE_KEY",
    "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InV1bWV6d293cnR1bWJvbnNvdHljIiwicm9sZSI6ImFub24iLCJpYXQiOjE3NjExNTk1MzUsImV4cCI6MjA3NjczNTUzNX0.dZGdfqa7BuYH6_W3yqirn8DsuEoffnyBm1qoLU-K0A0",
)
supabase = create_client(SUPABASE_URL, SUPABASE_KEY)

# ==========================================================
#           ENSURE last_edit TABLE EXISTS (attempt)
# ==========================================================
def ensure_last_edit_table():
    """
    Try a safe select; if it fails, attempt to create the last_edit table.
    This uses a Supabase RPC to execute raw SQL; depending on your project's
    RLS and RPC permissions this may or may not succeed — errors are shown
    to the app but the app continues to work.
    """
    try:
        supabase.table("last_edit").select("*").limit(1).execute()
    except Exception:
        try:
            # Some Supabase projects allow a SQL-executing RPC, others don't.
            # Try to create the minimal table. If it errors we'll show a warning and continue.
            supabase.rpc("execute_sql", {
                "sql": """
                CREATE TABLE IF NOT EXISTS public.last_edit (
                    id bigint PRIMARY KEY,
                    user_name text,
                    timestamp timestamptz DEFAULT now()
                );
                """
            }).execute()
        except Exception as e:
            # Not fatal — we will continue but warn the user so they can create the table manually.
            st.warning(f"No se pudo verificar/crear la tabla last_edit automáticamente: {e}")

ensure_last_edit_table()

# ==========================================================
#               PAGE CONFIGURATION
# ==========================================================
st.set_page_config(page_title="Valentina por Siempre", page_icon="VxS_logo.png", layout="wide")

# ==========================================================
#               LAST EDIT HELPERS
# ==========================================================
def update_last_edit(user_name):
    """Upsert the last_edit row (id=1). Called after add/edit/delete actions."""
    now = datetime.now().isoformat()
    try:
        supabase.table("last_edit").upsert({"id": 1, "user_name": user_name, "timestamp": now}).execute()
    except Exception as e:
        st.warning(f"⚠️ No se pudo actualizar el registro de edición en Supabase: {e}")

def get_last_edit():
    """Return tuple (user_name, timestamp) or (None, None)."""
    try:
        result = supabase.table("last_edit").select("*").eq("id", 1).execute()
        if result.data:
            record = result.data[0]
            return record.get("user_name"), record.get("timestamp")
    except Exception:
        return None, None
    return None, None

# ==========================================================
#                 ACCESS CONTROL (LOGIN)
# ==========================================================
if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
    st.session_state.user_name = ""

if not st.session_state.authenticated:
    st.sidebar.title("🔐 Ingreso")
    user_key = st.sidebar.text_input("Introduce tu clave de acceso:", type="password")

    AUTHORIZED_KEYS = {
        "equipo_vxs": None,         # regular users must type their name
        "valentina_master": "Andrea"  # master key: name auto-filled
    }

    if user_key not in AUTHORIZED_KEYS:
        st.warning("Por favor, introduce la clave de acceso para continuar.")
        st.stop()

    if AUTHORIZED_KEYS[user_key] is None:
        user_name = st.sidebar.text_input("Tu nombre (para registrar ediciones):")
        if not user_name:
            st.info("Por favor, escribe tu nombre para continuar.")
            st.stop()
    else:
        user_name = AUTHORIZED_KEYS[user_key]

    if st.sidebar.button("Ingresar"):
        st.session_state.authenticated = True
        st.session_state.user_name = user_name
        st.rerun()

# ==========================================================
#                 STYLES & LOGO + TABLE WRAP CSS
# ==========================================================
def load_logo_base64(path: str):
    file_path = Path(path)
    if file_path.exists():
        with open(file_path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    return None

logo_b64 = load_logo_base64("VxS_logo.png")
st.markdown(f"""
    <style>
    .main {{ background-color: #f7f5f2 !important; }}
    .custom-title {{
        text-align: center;
        color: #352208;
        font-size: 48px;
        font-weight: 800;
        margin-top: -5px;
        margin-bottom: 15px;
    }}
    .corner-image {{
        position: fixed; top: 80px; right: 25px;
        width: 90px; border-radius: 50%;
        box-shadow: 0px 4px 10px rgba(0,0,0,0.2);
        z-index: 999;
    }}
    .bottom-left {{
        position: fixed; bottom: 10px; left: 15px;
        color: #666; font-size: 14px;
        background-color: rgba(255,255,255,0.9);
        padding: 5px 10px; border-radius: 8px;
    }}

    /* Make dataframe cells wrap so text doesn't cut off */
    [data-testid="stDataFrame"] div[role="gridcell"] {{
        white-space: normal !important;
        overflow-wrap: anywhere !important;
        word-break: break-word !important;
    }}
    </style>
    {'<img src="data:image/png;base64,' + logo_b64 + '" class="corner-image">' if logo_b64 else ''}
""", unsafe_allow_html=True)

st.markdown("<h1 class='custom-title'>💛 Valentina por Siempre</h1>", unsafe_allow_html=True)

# ==========================================================
#                 HELPER FUNCTIONS
# ==========================================================
def calculate_age(dob):
    """Return integer age. dob may be datetime.date or string 'YYYY-MM-DD'."""
    if pd.isna(dob):
        return ""
    if isinstance(dob, str):
        try:
            dob = datetime.strptime(dob, "%Y-%m-%d").date()
        except Exception:
            dob = pd.to_datetime(dob, errors='coerce').date()
    today = date.today()
    return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))

def style_excel(df, filename):
    """
    Save the dataframe to excel, remove time from date columns,
    style header, freeze top row and highlight palliative rows.
    """
    df_copy = df.copy()
    for col in ["fecha_nacimiento", "fecha_ultimo_apoyo"]:
        if col in df_copy.columns:
            df_copy[col] = pd.to_datetime(df_copy[col], errors='coerce').dt.date

    df_copy.to_excel(filename, index=False)

    wb = load_workbook(filename)
    ws = wb.active

    # Column widths (adjust if you want different widths)
    col_widths = {
        1: 12, 2: 30, 3: 20, 4: 30, 5: 20,
        6: 20, 7: 25, 8: 25, 9: 25, 10: 25,
        11: 40, 12: 20, 13: 15, 14: 20, 15: 20
    }
    for i, width in col_widths.items():
        # Excel columns are A, B, C...; openpyxl expects letter names
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = width

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="FF8330")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Freeze header row
    ws.freeze_panes = "A2"

    # Highlight palliative rows (look up the column)
    paliativos_fill = PatternFill("solid", fgColor="FFAB66")
    paliativos_col = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "cuidados_paliativos":
            paliativos_col = idx
            break

    if paliativos_col:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            v = row[paliativos_col - 1].value
            if v in (1, True, "1", "true", "True", "TRUE"):
                for cell in row:
                    cell.fill = paliativos_fill

    wb.save(filename)

# ==========================================================
#                 MAIN INTERFACE
# ==========================================================
if st.session_state.authenticated:
    page = st.sidebar.radio(
        "Navegación",
        ["➕ Agregar Paciente", "🖊️ Editar Paciente", "📋 Ver Pacientes", "🎂 Cumpleaños"]
    )

    # ---------------- ADD PATIENT ----------------
    if page == "➕ Agregar Paciente":
        st.subheader("Ingresar nuevo paciente")
        with st.form("add_patient_form"):
            nombre = st.text_input("Nombre del paciente")
            fecha_nacimiento = st.date_input("Fecha de nacimiento", min_value=date(1900, 1, 1))
            nombre_tutor = st.text_input("Nombre del tutor")
            diagnostico = st.text_input("Diagnóstico")
            etapa_tratamiento = st.selectbox("Etapa del tratamiento", [
                "Diagnóstico inicial", "En tratamiento", "En vigilancia", "Cuidados paliativos"
            ])
            hospital = st.text_input("Hospital")
            estado_origen = st.text_input("Estado de origen")
            telefono_contacto = st.text_input("Celular de contacto")
            apoyos_entregados = st.text_input("Apoyos entregados")
            fecha_ultimo_apoyo = st.date_input("Fecha del último apoyo", value=None)
            notas = st.text_area("Notas")
            estado = st.selectbox("Estado del paciente", ["activo", "vigilancia", "fallecido"])
            cuidados_paliativos = st.checkbox("¿Está en cuidados paliativos?")

            submitted = st.form_submit_button("Agregar paciente")
            if submitted:
                payload = {
                    "nombre": nombre,
                    "fecha_nacimiento": str(fecha_nacimiento),
                    "nombre_tutor": nombre_tutor,
                    "diagnostico": diagnostico,
                    "etapa_tratamiento": etapa_tratamiento,
                    "hospital": hospital,
                    "estado_origen": estado_origen,
                    "telefono_contacto": telefono_contacto,
                    "apoyos_entregados": apoyos_entregados,
                    "fecha_ultimo_apoyo": str(fecha_ultimo_apoyo) if fecha_ultimo_apoyo else None,
                    "notas": notas,
                    "estado": estado,
                    "cuidados_paliativos": cuidados_paliativos
                }
                try:
                    supabase.table("pacientes").insert(payload).execute()
                    # update last edit AFTER insert
                    update_last_edit(st.session_state.user_name)
                    st.success(f"✅ Paciente agregado exitosamente por {st.session_state.user_name}.")
                    st.rerun()
                except Exception as e:
                    st.error(f"No se pudo agregar el paciente: {e}")

    # ---------------- EDIT PATIENT ----------------
    elif page == "🖊️ Editar Paciente":
        st.subheader("Editar información de paciente existente")
        # fetch all patients sorted by id so ordering is stable
        res = supabase.table("pacientes").select("*").order("id", count="exact").execute()
        df_full = pd.DataFrame(res.data)
        if df_full.empty:
            st.info("No hay pacientes para editar.")
        else:
            # present ids and names so user can pick the exact record
            df_full = df_full.sort_values("id", ascending=True).reset_index(drop=True)
            id_to_label = df_full.apply(lambda r: f"{r['id']} — {r['nombre']}", axis=1).tolist()
            selected_label = st.selectbox("Selecciona paciente (ID — Nombre):", id_to_label)
            selected_id = int(selected_label.split("—")[0].strip())

            patient = df_full[df_full["id"] == selected_id].iloc[0]

            with st.form("edit_patient_form"):
                nombre = st.text_input("Nombre del paciente", value=patient.get("nombre", ""))
                fecha_nacimiento_val = patient.get("fecha_nacimiento", None)
                try:
                    fecha_nacimiento_default = pd.to_datetime(fecha_nacimiento_val).date() if pd.notna(fecha_nacimiento_val) else date(2000,1,1)
                except Exception:
                    fecha_nacimiento_default = date(2000,1,1)
                fecha_nacimiento = st.date_input("Fecha de nacimiento", value=fecha_nacimiento_default, min_value=date(1900,1,1))
                nombre_tutor = st.text_input("Nombre del tutor", value=patient.get("nombre_tutor", ""))
                diagnostico = st.text_input("Diagnóstico", value=patient.get("diagnostico", ""))
                etapa_tratamiento = st.selectbox("Etapa del tratamiento", [
                    "Diagnóstico inicial", "En tratamiento", "En vigilancia", "Cuidados paliativos"
                ], index=["Diagnóstico inicial", "En tratamiento", "En vigilancia", "Cuidados paliativos"].index(patient.get("etapa_tratamiento") if patient.get("etapa_tratamiento") in ["Diagnóstico inicial","En tratamiento","En vigilancia","Cuidados paliativos"] else "Diagnóstico inicial"))
                hospital = st.text_input("Hospital", value=patient.get("hospital", ""))
                estado_origen = st.text_input("Estado de origen", value=patient.get("estado_origen", ""))
                telefono_contacto = st.text_input("Celular de contacto", value=patient.get("telefono_contacto", ""))
                apoyos_entregados = st.text_input("Apoyos entregados", value=patient.get("apoyos_entregados", ""))
                fecha_ultimo_apoyo_val = patient.get("fecha_ultimo_apoyo", None)
                try:
                    fecha_ultimo_default = pd.to_datetime(fecha_ultimo_apoyo_val).date() if pd.notna(fecha_ultimo_apoyo_val) else None
                except Exception:
                    fecha_ultimo_default = None
                fecha_ultimo_apoyo = st.date_input("Fecha del último apoyo", value=fecha_ultimo_default) if fecha_ultimo_default else st.date_input("Fecha del último apoyo", value=None)
                notas = st.text_area("Notas", value=patient.get("notas", ""))
                estado = st.selectbox("Estado del paciente", ["activo", "vigilancia", "fallecido"], index=["activo","vigilancia","fallecido"].index(patient.get("estado") if patient.get("estado") in ["activo","vigilancia","fallecido"] else "activo"))
                cuidados_paliativos = st.checkbox("¿Está en cuidados paliativos?", value=patient.get("cuidados_paliativos", False))

                submitted_edit = st.form_submit_button("💾 Guardar cambios")
                if submitted_edit:
                    update_payload = {
                        "nombre": nombre,
                        "fecha_nacimiento": str(fecha_nacimiento),
                        "nombre_tutor": nombre_tutor,
                        "diagnostico": diagnostico,
                        "etapa_tratamiento": etapa_tratamiento,
                        "hospital": hospital,
                        "estado_origen": estado_origen,
                        "telefono_contacto": telefono_contacto,
                        "apoyos_entregados": apoyos_entregados,
                        "fecha_ultimo_apoyo": str(fecha_ultimo_apoyo) if fecha_ultimo_apoyo else None,
                        "notas": notas,
                        "estado": estado,
                        "cuidados_paliativos": cuidados_paliativos
                    }
                    try:
                        supabase.table("pacientes").update(update_payload).eq("id", selected_id).execute()
                        update_last_edit(st.session_state.user_name)
                        st.success("✅ Cambios guardados correctamente.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"No se pudo guardar: {e}")

    # ---------------- VIEW PATIENTS ----------------
    elif page == "📋 Ver Pacientes":
        st.subheader("❤️‍🩹 Lista de pacientes")
        # checkboxes for filtering multiple estados
        st.markdown("**Filtrar por estado:**")
        col1, col2, col3 = st.columns(3)
        with col1:
            filtro_activo = st.checkbox("Activo", value=True)
        with col2:
            filtro_vigilancia = st.checkbox("Vigilancia", value=False)
        with col3:
            filtro_fallecido = st.checkbox("Fallecido", value=False)

        search = st.text_input("🔍 Buscar paciente por nombre o diagnóstico")

        selected_estados = []
        if filtro_activo: selected_estados.append("activo")
        if filtro_vigilancia: selected_estados.append("vigilancia")
        if filtro_fallecido: selected_estados.append("fallecido")

        if not selected_estados:
            st.info("Selecciona al menos un estado para mostrar los pacientes.")
            st.stop()

        # fetch rows for those states and sort by id ascending
        try:
            q = supabase.table("pacientes").select("*").in_("estado", selected_estados).order("id", {"ascending": True}).execute()
            df = pd.DataFrame(q.data)
        except Exception as e:
            st.error(f"No se pudo leer la tabla pacientes: {e}")
            df = pd.DataFrame()

        if not df.empty:
            # drop columns that are not for display (none are added in this version)
            df["fecha_nacimiento"] = pd.to_datetime(df["fecha_nacimiento"], errors="coerce").dt.date
            df["Edad"] = df["fecha_nacimiento"].apply(calculate_age)

            if search:
                df = df[df["nombre"].str.contains(search, case=False, na=False) | df["diagnostico"].str.contains(search, case=False, na=False)]

            def highlight_paliativos(row):
                color = "#FFAB66" if row.get("cuidados_paliativos") in [1, True, "1", "true", "True", "TRUE"] else ""
                return [f"background-color: {color}"] * len(row)

            st.dataframe(df.style.apply(highlight_paliativos, axis=1), use_container_width=True)

            # ---- Deletion by ID with confirmation ----
            st.markdown("### 🗑️ Eliminar paciente por ID")
            ids = df["id"].tolist()
            selected_id_for_delete = st.selectbox("Seleccionar ID para eliminar:", options=ids)
            confirm_key = f"confirm_delete_{selected_id_for_delete}"
            if st.button("Eliminar paciente seleccionado"):
                # show confirmation prompt
                st.session_state[confirm_key] = True

            if st.session_state.get(confirm_key, False):
                st.warning(f"⚠️ ¿Seguro que quieres eliminar al paciente con ID {selected_id_for_delete}? Esta acción es irreversible.")
                c1, c2 = st.columns(2)
                with c1:
                    if st.button("✅ Sí, eliminar"):
                        try:
                            supabase.table("pacientes").delete().eq("id", int(selected_id_for_delete)).execute()
                            update_last_edit(st.session_state.user_name)
                            st.success(f"✅ Paciente con ID {selected_id_for_delete} eliminado correctamente.")
                            st.session_state.pop(confirm_key, None)
                            st.rerun()
                        except Exception as e:
                            st.error(f"No se pudo eliminar: {e}")
                with c2:
                    if st.button("❌ Cancelar"):
                        st.session_state.pop(confirm_key, None)
                        st.info("Operación cancelada.")

            # ---- Export to Excel ----
            if st.button("📥 Exportar a Excel"):
                filename = "pacientes_valentina.xlsx"
                try:
                    style_excel(df, filename)
                    with open(filename, "rb") as file:
                        st.download_button("Descargar archivo Excel", data=file, file_name=filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                except Exception as e:
                    st.error(f"No se pudo generar el Excel: {e}")

        else:
            st.info("No hay pacientes registrados con ese estado.")

    # ---------------- BIRTHDAYS ----------------
    elif page == "🎂 Cumpleaños":
        st.subheader("Cumpleaños del mes y del próximo mes")
        MONTHS_ES = {
            "January": "enero", "February": "febrero", "March": "marzo",
            "April": "abril", "May": "mayo", "June": "junio",
            "July": "julio", "August": "agosto", "September": "septiembre",
            "October": "octubre", "November": "noviembre", "December": "diciembre"
        }
        try:
            q = supabase.table("pacientes").select("*").neq("estado", "fallecido").order("id", {"ascending": True}).execute()
            df = pd.DataFrame(q.data)
        except Exception as e:
            st.error(f"No se pudo leer la tabla pacientes: {e}")
            df = pd.DataFrame()

        if not df.empty:
            df["fecha_nacimiento"] = pd.to_datetime(df["fecha_nacimiento"], errors="coerce").dt.date
            df["Edad"] = df["fecha_nacimiento"].apply(calculate_age)
            current_month = datetime.today().month
            next_month = (current_month % 12) + 1
            df_this_month = df[pd.to_datetime(df["fecha_nacimiento"]).dt.month == current_month]
            df_next_month = df[pd.to_datetime(df["fecha_nacimiento"]).dt.month == next_month]

            current_month_name = MONTHS_ES[datetime.today().strftime('%B')]
            st.markdown(f"### 🎉 Cumpleaños de **{current_month_name}**")
            if not df_this_month.empty:
                st.dataframe(df_this_month[["nombre", "fecha_nacimiento", "Edad", "estado"]])
            else:
                st.info("No hay cumpleaños este mes.")

            next_month_name_en = datetime(datetime.today().year, next_month, 1).strftime('%B')
            next_month_name = MONTHS_ES[next_month_name_en]
            st.markdown(f"### 🎈 Cumpleaños de **{next_month_name}**")
            if not df_next_month.empty:
                st.dataframe(df_next_month[["nombre", "fecha_nacimiento", "Edad", "estado"]])
            else:
                st.info("No hay cumpleaños el próximo mes.")
        else:
            st.info("No hay pacientes registrados.")

    # ---------------- FOOTER: last edit ----------------
    last_user, last_time = get_last_edit()
    if last_user and last_time:
        try:
            formatted_time = datetime.fromisoformat(last_time).strftime('%d/%m/%Y %H:%M')
        except Exception:
            formatted_time = last_time
        st.sidebar.markdown(
            f"<div class='bottom-left'>Última edición por <b>{last_user}</b> el {formatted_time}</div>",
            unsafe_allow_html=True
        )
